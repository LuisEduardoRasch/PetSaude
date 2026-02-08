"""Ferramentas para conversão e consolidação de altas em CSV.

Este módulo converte todos os arquivos ``.ods`` encontrados em uma pasta para
arquivos ``.csv``, concatena todos os CSVs em um único conjunto de dados,
remove duplicados com base em nome e data de alta e realiza uma separação
especial pela coluna ``Tipo de Alta``.

Fluxo principal (``main``)
--------------------------
- Converte arquivos ``.ods`` em CSVs temporários.
- Lê todos os CSVs (temporários + existentes) e concatena em um único
    :class:`pandas.DataFrame`.
- Remove registros duplicados usando, quando possível, as colunas
    "Pacientes"/"Nome" e "Dia Alta"/"Data" (comparação case-insensitive).
- Separa, em um CSV à parte, os registros cujo ``Tipo de Alta`` não seja
    "Melhorado" nem "Melhorada" (ignorando maiúsculas/minúsculas).
- Remove, do CSV principal consolidado, as colunas ``Tipo de Alta``,
    ``Telefone`` e ``Dia Alta``.
 - Efetua uma limpeza geral (remoção de linhas vazias e sem paciente) e grava
        o resultado final em ``merged_principal.xlsx`` e as altas não
        "Melhorado/Melhorada" em ``altas_nao_melhorado.xlsx``.

Uso
---
        python convert_merge_split.py --input-dir <pasta> --output-dir <pasta_destino>

As opções são todas opcionais; na ausência de parâmetros, o script utiliza a
pasta ``Arquivos`` como entrada e ``output`` como pasta de saída.
"""
import argparse
import csv
import json
import logging
import os
import re
import sys
import unicodedata
from pathlib import Path
from typing import Optional
from openpyxl.utils import get_column_letter

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

try:
    import ezodf
except Exception:
    ezodf = None

try:
    import pandas as pd
except Exception:
    pd = None

from clean_csv import clean_dataframe

logger = logging.getLogger(__name__)

DEFAULT_CONFIG = {
    "input_dir": "Arquivos",
    "output_dir": "output",
    "temp_dir": None,
    "log_level": "INFO",
    "required_columns": {
        "Pacientes": ["Pacientes", "Paciente", "Nome do Paciente"],
        "Tipo de Alta": ["Tipo de Alta", "Alta"],
        "Encaminhado": ["Encaminhado", "Encaminhamento"],
    },
}


def setup_logging(level: str) -> None:
    logging.basicConfig(
        level=getattr(logging, level.upper(), logging.INFO),
        format="%(asctime)s [%(levelname)s] %(message)s",
    )


def load_config(config_path: Optional[Path], script_dir: Path) -> dict:
    config = DEFAULT_CONFIG.copy()
    config["required_columns"] = DEFAULT_CONFIG["required_columns"].copy()

    if config_path is None:
        default_path = script_dir / "config.json"
        if default_path.exists():
            config_path = default_path

    if config_path is not None and config_path.exists():
        with open(config_path, "r", encoding="utf-8") as f:
            loaded = json.load(f)

        config.update({k: v for k, v in loaded.items() if k != "required_columns"})
        if "required_columns" in loaded:
            config["required_columns"].update(loaded["required_columns"])

    return config


def validate_schema(df, required_columns: dict, context: str) -> dict:
    missing = []
    resolved = {}

    for logical_name, candidates in required_columns.items():
        if not isinstance(candidates, list):
            candidates = [candidates]

        col = find_column(df, candidates)
        if col is None:
            missing.append(logical_name)
        else:
            resolved[logical_name] = col

    if missing:
        raise ValueError(f"Schema inválido em {context}. Faltando: {', '.join(missing)}")

    return resolved

def ensure_dependencies():
    missing = []
    if ezodf is None:
        missing.append('ezodf')
    if pd is None:
        missing.append('pandas')
    if missing:
        print('Dependências ausentes: %s' % ', '.join(missing))
        print('Instale com: pip install -r requirements.txt')
        sys.exit(1)

    # Ajuste defensivo: algumas versões do ezodf não possuem a chave
    # '.ods' no mapa interno de MIME types, o que causa KeyError.
    try:
        import ezodf.document as _ezdoc  # type: ignore[import]

        mts = getattr(_ezdoc, 'MIMETYPES', None)
        if isinstance(mts, dict):
            # Se existir a chave 'ods' mas não '.ods', cria um alias.
            if 'ods' in mts and '.ods' not in mts:
                mts['.ods'] = mts['ods']
            if 'ods' in mts and '.ODS' not in mts:
                mts['.ODS'] = mts['ods']
    except Exception:
        # Se o patch falhar por qualquer motivo, seguimos usando o padrão.
        pass


def ods_to_csv(ods_path: Path, out_dir: Path):
    """Converte um arquivo .ods para um ou mais CSVs (uma por planilha)."""
    # Garante que o caminho seja passado como string para o ezodf.
    doc = ezodf.opendoc(str(ods_path))
    created = []
    for sheet in doc.sheets:
        if sheet.nrows() == 0:
            continue
            
        # Encontra o cabeçalho
        header_row = None
        data_start_row = 0
        
        for r in range(min(5, sheet.nrows())):  # Procura nas primeiras 5 linhas
            row_data = []
            for c in range(sheet.ncols()):
                cell = sheet[r, c]
                val = cell.value
                if val is not None:
                    row_data.append(str(val).strip().lower())
                else:
                    row_data.append('')
            
            # Verifica se esta linha parece ser um cabeçalho
            if any('paciente' in cell or 'nome' in cell for cell in row_data):
                header_row = r
                data_start_row = r + 1
                break
        
        # Extrai dados
        all_rows = []
        
        # Adiciona cabeçalho padronizado
        standard_header = ['Pacientes', 'Tipo de Alta', 'Telefone', 'Dia Alta', 'Cid', 'Endereço', 'Encaminhado']
        all_rows.append(standard_header)
        
        for r in range(data_start_row, sheet.nrows()):
            row = []
            for c in range(sheet.ncols()):
                cell = sheet[r, c]
                val = cell.value
                if val is None:
                    row.append('')
                else:
                    # Format dates properly
                    if hasattr(val, 'strftime'):
                        row.append(val.strftime('%Y-%m-%d'))
                    else:
                        row.append(str(val).strip())
            
            # Processa a linha e pode gerar múltiplas linhas
            processed_rows = process_data_row(row)
            for processed_row in processed_rows:
                if processed_row and any(cell.strip() for cell in processed_row):
                    # Padroniza para 7 colunas
                    while len(processed_row) < len(standard_header):
                        processed_row.append('')
                    processed_row = processed_row[:len(standard_header)]
                    
                    # Só adiciona se tem nome de paciente
                    if processed_row[0].strip():
                        all_rows.append(processed_row)
        
        # Only save if there are meaningful rows
        if len(all_rows) > 1:  # At least header + 1 data row
            # Determine filename - só salva a primeira planilha (Plan1)
            safe_sheet = ''.join(ch if ch.isalnum() or ch in (' ', '_', '-') else '_' for ch in sheet.name)
            if 'plan1' in safe_sheet.lower() or sheet == doc.sheets[0]:
                out_name = ods_path.stem + '__' + safe_sheet + '.csv'
                out_path = out_dir / out_name
                with out_path.open('w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    for r in all_rows:
                        writer.writerow(r)
                created.append(out_path)
    return created


def process_data_row(row):
    """Processa uma linha de dados e pode retornar múltiplas linhas se houver dados misturados."""
    if not row or not any(cell.strip() for cell in row):
        return []
    
    # Remove colunas vazias do final
    while row and not row[-1].strip():
        row.pop()
    
    if not row:
        return []
    
    # Se a primeira coluna está vazia, remove
    if len(row) >= 2 and not row[0].strip():
        row = row[1:]
    
    if not row or not row[0].strip():
        return []
    
    # Detecta padrão onde o segundo nome está na coluna "Tipo de Alta"
    # Isso acontece quando há dois pacientes em uma linha
    if (
        len(row) >= 2
        and row[1].strip()
        and row[1].strip().upper() not in [
            'MELHORADA',
            'MELHORADO',
            'ALTA',
            'OBITO',
            'TRANSFERENCIA',
            'ABANDONO',
        ]
    ):
        
        # O segundo campo parece ser um nome, não um tipo de alta
        second_name = row[1].strip()
        
        # Verifica se é realmente um nome (só letras e espaços)
        if all(c.isalpha() or c.isspace() for c in second_name) and len(second_name.split()) >= 2:
            # Temos dois pacientes nesta linha
            first_name = row[0].strip()
            
            # IMPORTANTE: O segundo nome é que deve ser considerado o paciente principal
            # O primeiro nome parece ser um registro anterior ou erro de formatação
            
            # Cria linha para o segundo nome (paciente principal) com os dados completos
            row_main = [second_name]
            if len(row) > 2:
                row_main.extend(row[2:])  # Use os dados reais (tipo alta, telefone, etc.)
            else:
                row_main.extend([''] * 6)  # Preenche campos vazios
            
            # OPCIONAL: Também criar linha para o primeiro nome, mas sem dados completos
            # (pode ser um paciente adicional ou erro - deixamos para investigação manual)
            row_additional = [first_name, '', '', '', '', '', '']
            
            # Retorna o paciente principal primeiro (segundo nome) e depois o adicional
            return [row_main, row_additional]
    
    # Verifica se o primeiro campo tem múltiplos nomes separados por vírgula
    first_field = row[0].strip()
    
    if ',' in first_field:
        parts = first_field.split(',')
        if len(parts) == 2:
            # Caso: "NOME1,NOME2" 
            name1 = parts[0].strip()
            name2 = parts[1].strip()
            
            if (name1 and name2 and 
                all(c.isalpha() or c.isspace() for c in name1) and
                all(c.isalpha() or c.isspace() for c in name2) and
                len(name1.split()) >= 2 and len(name2.split()) >= 2):
                
                # Cria primeira linha
                row1 = [name1] + row[1:]
                # Cria segunda linha com o mesmo dados mas nome diferente
                row2 = [name2] + row[1:]
                return [row1, row2]
    
    # Caso de nomes muito longos - pode ser dois nomes juntos
    elif len(first_field.split()) > 6:
        words = first_field.split()
        # Tenta identificar onde um nome termina e outro começa
        mid_point = len(words) // 2
        name1 = ' '.join(words[:mid_point])
        name2 = ' '.join(words[mid_point:])
        
        # Verifica se ambos parecem nomes válidos
        if (name1 and name2 and 
            len(name1.split()) >= 2 and len(name2.split()) >= 2 and
            all(word[0].isupper() for word in name1.split() if word) and
            all(word[0].isupper() for word in name2.split() if word)):
            
            row1 = [name1] + row[1:]
            row2 = [name2] + row[1:]
            return [row1, row2]
    
    # Caso normal - retorna linha única
    return [row]



def find_files(input_dir: Path):
    ods = list(input_dir.rglob('*.ods'))
    csvs = list(input_dir.rglob('*.csv'))
    return ods, csvs


def concat_csvs(csv_paths, out_path: Path):
    """Concatena múltiplos arquivos CSV em um único :class:`pandas.DataFrame`.

    A função procura normalizar nomes de colunas vindos de fontes diversas
    para um conjunto padrão ("Pacientes", "Tipo de Alta", "Telefone",
    "Dia Alta", "Cid", "Endereço", "Encaminhado"), limpa espaços em
    branco, corrige alguns casos comuns de desalinhamento de dados e remove
    linhas sem nome de paciente.

    Parameters
    ----------
    csv_paths : iterable of pathlib.Path
        Caminhos dos arquivos CSV a serem concatenados.
    out_path : pathlib.Path
        Caminho de saída utilizado apenas se nenhum DataFrame válido for
        gerado; nesse caso, um arquivo vazio é criado nesse local.

    Returns
    -------
    pandas.DataFrame or pathlib.Path
        DataFrame concatenado com todas as linhas válidas ou, se nenhum
        CSV válido for encontrado, o próprio ``out_path``.
    """
    dfs = []

    for p in csv_paths:
        try:
            df = pd.read_csv(p, dtype=str)
            if df.empty:
                continue

            # Limpa nomes de colunas e tenta mapeá-los para o padrão interno
            df.columns = [str(c).strip() for c in df.columns]

            col_map = {}

            col = find_column(df, ['pacientes', 'nome'])
            if col and col != 'Pacientes':
                col_map[col] = 'Pacientes'

            col = find_column(df, ['tipo de alta', 'tipo_alta'])
            if col and col != 'Tipo de Alta':
                col_map[col] = 'Tipo de Alta'

            col = find_column(df, ['telefone', 'fone'])
            if col and col != 'Telefone':
                col_map[col] = 'Telefone'

            col = find_column(df, ['dia alta', 'data', 'dia de alta'])
            if col and col != 'Dia Alta':
                col_map[col] = 'Dia Alta'

            col = find_column(df, ['cid'])
            if col and col != 'Cid':
                col_map[col] = 'Cid'

            col = find_column(df, ['endereço', 'endereco'])
            if col and col != 'Endereço':
                col_map[col] = 'Endereço'

            col = find_column(df, ['encaminhado'])
            if col and col != 'Encaminhado':
                col_map[col] = 'Encaminhado'

            if col_map:
                df = df.rename(columns=col_map)

            # Limpa dados: converte tudo para string e remove espaços
            for col in df.columns:
                df[col] = df[col].fillna('').astype(str).str.strip()

            # Corrige linhas em que Encaminhado está vazio e Endereço parece CAPS/UBS/HOSPITAL
            enc_col = 'Encaminhado' if 'Encaminhado' in df.columns else find_column(df, ['encaminhado'])
            end_col = 'Endereço' if 'Endereço' in df.columns else find_column(df, ['endereço', 'endereco'])

            if enc_col and end_col:
                for idx in df.index:
                    end_val = str(df.loc[idx, end_col]).upper()
                    if df.loc[idx, enc_col] == '' and end_val.startswith(('CAPS', 'UBS', 'HOSPITAL')):
                        df.loc[idx, enc_col] = df.loc[idx, end_col]
                        df.loc[idx, end_col] = ''

            # Remove linhas sem paciente (prováveis cabeçalhos/junk)
            pac_col = 'Pacientes' if 'Pacientes' in df.columns else find_column(df, ['pacientes', 'nome'])
            if pac_col:
                df = df[df[pac_col].fillna('').astype(str).str.strip() != '']

            # Remove linhas completamente vazias
            df = df.dropna(how='all')

            if not df.empty:
                dfs.append(df)
                logger.info('Processado %s: %s linhas válidas', p, len(df))

        except Exception as e:
            logger.warning('Erro lendo %s: %s', p, e)

    if not dfs:
        # Cria arquivo vazio se nada foi lido
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text('')
        return out_path

    big = pd.concat(dfs, ignore_index=True, sort=False)

    # Limpeza final global
    big = big.dropna(how='all')
    if 'Pacientes' in big.columns:
        big['Pacientes'] = big['Pacientes'].fillna('').astype(str)
        big = big[big['Pacientes'].str.strip() != '']

    return big


def find_column(df, candidates):
    """Return the first column name in df that matches any of the candidates (case-insensitive)."""
    cols = list(df.columns)
    for cand in candidates:
        for c in cols:
            if c and str(c).strip().lower() == cand.lower():
                return c
    # try more flexible contains match
    for c in cols:
        for cand in candidates:
            if cand.lower() in str(c).strip().lower():
                return c
    return None


def _normalize_cell_for_empty(value):
    """Normaliza um valor de célula para detecção de vazio.

    Converte para string, remove espaços e aspas nas pontas e considera
    como vazio também valores compostos apenas por vírgulas (por exemplo
    ",,,"), que surgem de linhas com apenas separadores no CSV original.
    """
    if value is None:
        s = ''
    else:
        s = str(value)

    s = s.strip().strip('"').strip("'")

    if not s:
        return ''

    # Se após limpeza o valor é apenas um monte de vírgulas, trata como vazio
    if all(ch == ',' for ch in s):
        return ''

    return s

def _normalize_text_key(value):
    """Normaliza texto para comparação de duplicatas.

    - Converte para string.
    - Remove espaços extras nas pontas.
    - Remove acentos (ex.: 'Á', 'ã', 'Ç' -> 'A', 'A', 'C').
    - Converte "ç"/"Ç" em "c"/"C".
    - Colapsa múltiplos espaços internos em um único espaço.
    - Converte para maiúsculas.

    Retorna string vazia para valores nulos ou em branco após limpeza.
    """
    if value is None:
        return ''

    s = str(value)

    # Normaliza espaços (inclui espaços não separáveis)
    s = s.replace('\u00A0', ' ')
    s = s.strip()
    if not s:
        return ''

    # Remove acentos usando decomposição Unicode
    s = unicodedata.normalize('NFD', s)
    s = ''.join(ch for ch in s if not unicodedata.combining(ch))

    # Garantia extra para ç/Ç (apesar de já coberto acima)
    s = s.replace('ç', 'c').replace('Ç', 'C')

    # Colapsa múltiplos espaços internos
    s = re.sub(r'\s+', ' ', s)

    # Normaliza para maiúsculas
    s = s.upper()

    return s


def remove_duplicates(df):
    """Remove duplicatas de pacientes utilizando nome + endereço (preferencialmente).

    A deduplicação segue a seguinte prioridade:

    1. Se existirem colunas que representem nome do paciente e endereço,
       remove duplicatas por (nome_normalizado, endereco_normalizado).
    2. Caso não exista coluna de endereço, tenta remover duplicatas por
       (nome_normalizado, data_normalizada), usando as colunas de dia de alta/data.
    3. Se nada disso estiver disponível, remove duplicatas considerando
       a linha completa.

    A normalização converte valores para uma forma canônica, removendo
    acentos, transformando "ç" em "c", colapsando espaços duplos e
    aplicando maiúsculas, para evitar que pequenas variações de escrita
    permitam que duplicatas passem.
    """
    name_col = find_column(df, ['pacientes', 'nome'])
    addr_col = find_column(df, ['endereço', 'endereco'])
    date_col = find_column(df, ['dia alta', 'data'])

    # Nenhuma coluna de nome encontrada: cai para duplicata de linha inteira
    if not name_col:
        logger.warning('Coluna de nome de paciente não encontrada; removendo duplicados completos.')
        return df.drop_duplicates()

    # Cria colunas normalizadas auxiliares (não são gravadas no resultado final)
    df = df.copy()

    df['_NORM_NAME'] = df[name_col].apply(_normalize_text_key)

    if addr_col:
        df['_NORM_ADDR'] = df[addr_col].apply(_normalize_text_key)

    if date_col:
        df['_NORM_DATE'] = df[date_col].apply(_normalize_text_key)

    # 1) Tenta nome + endereço
    if addr_col:
        logger.info('Removendo duplicados por colunas: %s + %s (normalizados)', name_col, addr_col)
        deduped = df.drop_duplicates(subset=['_NORM_NAME', '_NORM_ADDR'], keep='first')
    # 2) Senão, tenta nome + data
    elif date_col:
        logger.info('Removendo duplicados por colunas: %s + %s (normalizados)', name_col, date_col)
        deduped = df.drop_duplicates(subset=['_NORM_NAME', '_NORM_DATE'], keep='first')
    # 3) Fallback: linha inteira
    else:
        logger.info('Sem endereço/data; removendo duplicados por linha completa.')
        deduped = df.drop_duplicates(keep='first')

    # Remove colunas auxiliares antes de devolver
    for aux in ['_NORM_NAME', '_NORM_ADDR', '_NORM_DATE']:
        if aux in deduped.columns:
            deduped = deduped.drop(columns=[aux])

    return deduped


def split_by_tipo_alta(df, output_dir: Path):
    """
    Separa registros com tipo de alta diferente de 'Melhorado' ou 'Melhorada'.

    Parameters
    ----------
    df : pandas.DataFrame
        DataFrame consolidado contendo ao menos a coluna 'Tipo de Alta'.
    output_dir : pathlib.Path
        Diretório onde o CSV com tipos de alta não padrão será salvo.

    Returns
    -------
    main_df : pandas.DataFrame
        DataFrame contendo apenas altas 'Melhorado'/'Melhorada' ou sem
        informação de tipo de alta (coluna vazia).
    other_file : pathlib.Path or None
        Caminho do CSV gerado com tipos de alta não padrão, ou None
        caso nenhum registro tenha sido separado.
    """
    tipo_col = find_column(df, ['tipo de alta'])
    if tipo_col is None:
        logger.warning('Coluna "Tipo de Alta" não encontrada; nenhum arquivo extra será gerado.')
        return df, None

    series = df[tipo_col].fillna('').astype(str).str.strip()
    norm = series.str.lower()

    # Registros considerados "não padrão":
    # - tipo de alta não vazio
    # - e diferente de "melhorado" ou "melhorada"
    mask_other = (series != '') & ~norm.isin(['melhorado', 'melhorada'])

    if not mask_other.any():
        logger.info('Nenhum registro com Tipo de Alta diferente de "Melhorado/Melhorada" encontrado.')
        return df, None

    other_df = df[mask_other].copy()
    main_df = df[~mask_other].copy()

    output_dir.mkdir(parents=True, exist_ok=True)
    other_file = output_dir / 'altas_nao_melhorado.xlsx'

    # Grava arquivo de altas não "Melhorado/Melhorada" com ajuste básico de colunas
    with pd.ExcelWriter(other_file, engine='openpyxl') as writer:
        other_df.to_excel(writer, index=False, sheet_name='Altas')
        ws = writer.sheets['Altas']

        # Autoajuste simples de largura das colunas com base no conteúdo
        for col_idx, col_name in enumerate(other_df.columns, start=1):
            col_len = max(
                other_df[col_name].astype(str).map(len).max() if not other_df.empty else 0,
                len(col_name),
            )
            col_letter = get_column_letter(col_idx)
            # Largura em unidades de caractere aproximada
            ws.column_dimensions[col_letter].width = min(col_len + 2, 80)

        # Exemplos (comente/descomente para ajustes específicos):
        # ws.column_dimensions['A'].width = 40   # primeira coluna mais larga
        # ws.row_dimensions[1].height = 18      # cabeçalho mais alto (em pontos aprox.)

    logger.info(
        'Registros com Tipo de Alta diferente de "Melhorado/Melhorada" '
        'salvos em: %s (%s linhas)',
        other_file,
        len(other_df),
    )

    return main_df, other_file


def split_by_encaminhado(df, output_dir: Path):
    # find encaminhado column case-insensitive
    enc_col = find_column(df, ['encaminhado'])
    
    if enc_col is None:
        # write all to single file
        out = output_dir / 'all_encaminhado_missing.csv'
        output_dir.mkdir(parents=True, exist_ok=True)
        df.to_csv(out, index=False)
        logger.warning('Coluna encaminhado não encontrada. Arquivo escrito em: %s', out)
        return [out]
    
    # Make a copy to avoid the warning
    df_work = df.copy()
    
    # Clean and normalize the encaminhado values
    df_work[enc_col] = df_work[enc_col].str.strip().str.upper()
    
    # Group similar values (e.g., "CAPS TRÊS VENDAS" and "CAPS TRES VENDAS")
    def normalize_encaminhado(val):
        if pd.isna(val) or val == '':
            return 'VAZIO'
        val = str(val).strip().upper()
        # Normalize common variations
        val = val.replace('TRÊS', 'TRES')
        val = val.replace('  ', ' ')  # Remove double spaces
        return val
    
    df_work[enc_col] = df_work[enc_col].apply(normalize_encaminhado)
    
    files = []
    output_dir.mkdir(parents=True, exist_ok=True)
    
    for val, group in df_work.groupby(enc_col):
        safe = ''.join(ch if ch.isalnum() or ch in (' ', '_', '-') else '_' for ch in str(val))
        fname = f'encaminhado__{safe or "vazio"}.csv'
        out = output_dir / fname
        group.to_csv(out, index=False)
        files.append(out)
        logger.info('Arquivo criado: %s com %s registros', fname, len(group))
    
    return files


def main():
    """
    Executa o fluxo completo de processamento de arquivos ODS/CSV.

    Etapas
    ------
    1. Localiza arquivos de entrada.
    2. Converte arquivos ODS em CSV temporários.
    3. Concatena todos os CSVs em um único DataFrame.
    4. Remove registros duplicados.
    5. Separa registros cujo 'Tipo de Alta' é diferente de
       'Melhorado' ou 'Melhorada' em um CSV à parte.
    6. Remove as colunas 'Tipo de Alta', 'Telefone' e 'Dia Alta'
       do CSV principal.
    7. Remove linhas vazias e grava o resultado final em disco.
    """
    parser = argparse.ArgumentParser(
        description=(
            'Converter .ods→.csv, concatenar, deduplicar e separar por Tipo de Alta '
            '(Melhorado/Melhorada vs demais).'
        )
    )
    parser.add_argument(
        '--input-dir', '-i', default=None,
        help='Pasta com arquivos .ods/.csv (padrão: ./Arquivos)',
    )
    parser.add_argument(
        '--output-dir', '-o', default=None,
        help='Pasta de saída para arquivos resultantes (padrão: ./output)',
    )
    parser.add_argument(
        '--temp-dir', '-t', default=None,
        help='Pasta temporária para CSVs convertidos (padrão: output-dir/temp_csvs)',
    )
    parser.add_argument(
        '--config', '-c', default=None,
        help='Arquivo de configuração JSON (padrão: ./config.json)',
    )
    args = parser.parse_args()

    # Diretórios padrão: usa 'Arquivos' como entrada e 'output' como saída.
    script_dir = Path(__file__).resolve().parent
    config_path = Path(args.config).resolve() if args.config else None
    config = load_config(config_path, script_dir)
    setup_logging(config.get("log_level", "INFO"))

    input_dir = Path(args.input_dir).resolve() if args.input_dir else (script_dir / config["input_dir"]).resolve()
    output_dir = Path(args.output_dir).resolve() if args.output_dir else (script_dir / config["output_dir"]).resolve()
    temp_dir = Path(args.temp_dir).resolve() if args.temp_dir else (
        output_dir / "temp_csvs" if config.get("temp_dir") is None else Path(config["temp_dir"]).resolve()
    )

    if args.input_dir is None:
        logger.info('Nenhum --input-dir informado; usando padrão: %s', input_dir)
    if args.output_dir is None:
        logger.info('Nenhum --output-dir informado; usando padrão: %s', output_dir)

    ensure_dependencies()

    ods_files, csv_files = find_files(input_dir)
    logger.info('Encontrado %s .ods e %s .csv em %s', len(ods_files), len(csv_files), input_dir)

    temp_dir.mkdir(parents=True, exist_ok=True)

    # 1) Converte ODS para CSV temporários
    for ods in ods_files:
        created = ods_to_csv(ods, temp_dir)
        logger.info('Convertido %s -> %s CSV(s)', ods, len(created))

    # 2) Coleta todos os CSVs (temporários + já existentes)
    all_csvs = list(temp_dir.rglob('*.csv')) + csv_files
    logger.info('Total CSVs para concatenar: %s', len(all_csvs))

    # 3) Concatena todos os CSVs em um único DataFrame
    big = concat_csvs(all_csvs, output_dir / 'merged.csv')
    if isinstance(big, Path):
        logger.warning('Nenhum CSV válido para concatenar; saindo')
        return

    validate_schema(big, config.get("required_columns", {}), "CSV concatenado")

    # 4) Remove duplicados
    deduped = remove_duplicates(big)

    # 5) Separa registros por Tipo de Alta
    deduped, other_file = split_by_tipo_alta(deduped, output_dir)

    # 6) Remove colunas não desejadas do CSV principal
    cols_to_drop = [col for col in ['Tipo de Alta', 'Telefone'] if col in deduped.columns]
    if cols_to_drop:
        logger.info('Removendo colunas do CSV principal: %s', ", ".join(cols_to_drop))
        deduped = deduped.drop(columns=cols_to_drop)

    # 7) Limpeza final: remove linhas vazias ou apenas com vírgulas/aspas
    # Cria uma cópia normalizada para avaliar quais linhas têm conteúdo real.
    cleaned = deduped.copy()
    for col in cleaned.columns:
        cleaned[col] = cleaned[col].map(_normalize_cell_for_empty)

    # Remove linhas em que todas as colunas ficam vazias após normalização
    mask_any_value = (cleaned != '').any(axis=1)
    deduped = deduped[mask_any_value]
    cleaned = cleaned[mask_any_value]

    # Remove linhas em que a coluna de paciente está vazia (ou só vírgulas/aspas)
    if 'Pacientes' in cleaned.columns:
        mask_paciente = cleaned['Pacientes'] != ''
        deduped = deduped[mask_paciente]
        cleaned = cleaned[mask_paciente]

        # 8) Deduplicação FINAL: mesmo paciente + mesmo endereço (normalizados)
    if 'Pacientes' in cleaned.columns and 'Endereço' in cleaned.columns:
        norm_name = cleaned['Pacientes'].map(_normalize_text_key)
        norm_addr = cleaned['Endereço'].map(_normalize_text_key)

        mask_with_addr = (norm_name != '') & (norm_addr != '')
        idx_with_addr = cleaned.index[mask_with_addr]

        if len(idx_with_addr) > 0:
            keys = norm_name[mask_with_addr] + '|' + norm_addr[mask_with_addr]
            duplicated = keys.duplicated(keep='first')
            drop_idx = idx_with_addr[duplicated]
            if len(drop_idx) > 0:
                logger.info('Removendo duplicatas finais por Pacientes+Endereço: %s linhas', len(drop_idx))
                deduped = deduped.drop(index=drop_idx)

    # Normaliza datas e ordena por data crescente antes de salvar
    date_col = find_column(deduped, ['dia alta', 'data'])
    if date_col:
        parsed_dates = pd.to_datetime(deduped[date_col], errors='coerce', dayfirst=True)
        deduped[date_col] = parsed_dates.dt.strftime('%Y-%m-%d')
        deduped['_SORT_DATE'] = parsed_dates
        deduped = deduped.sort_values(by='_SORT_DATE', na_position='last')
        deduped = deduped.drop(columns=['_SORT_DATE'])
        logger.info('Datas normalizadas em %s e ordenação aplicada.', date_col)

    # Reorganiza o índice antes de salvar
    deduped = deduped.reset_index(drop=True)

    # Grava arquivo principal final em formato Excel (.xlsx) com ajuste básico de colunas
    output_dir.mkdir(parents=True, exist_ok=True)
    merged_out = output_dir / 'merged_principal.xlsx'

    with pd.ExcelWriter(merged_out, engine='openpyxl') as writer:
        deduped.to_excel(writer, index=False, sheet_name='Altas')
        ws = writer.sheets['Altas']

        # Autoajuste simples de largura das colunas com base no conteúdo
        for col_idx, col_name in enumerate(deduped.columns, start=1):
            col_len = max(
                deduped[col_name].astype(str).map(len).max() if not deduped.empty else 0,
                len(col_name),
            )
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min(col_len + 2, 80)

        # Exemplos opcionais de ajuste fino:
        # ws.column_dimensions['A'].width = 40   # "Pacientes" mais larga
        # ws.row_dimensions[1].height = 18      # cabeçalho mais alto

    logger.info('Arquivo principal escrito em: %s', merged_out)
    if other_file is not None:
        logger.info('CSV com tipos de alta não "Melhorado/Melhorada": %s', other_file)

    generate_charts(deduped, output_dir)

def generate_patient_count_report(clean_dir: Path, report_file: Path):
    """Gera relatório com quantidade de pacientes por arquivo CAPS."""
    if not clean_dir.exists():
        logger.warning('Pasta %s não encontrada', clean_dir)
        return
    
    # Collect data for all files
    caps_data = []
    total_patients = 0
    
    for csv_file in sorted(clean_dir.glob('*.csv')):
        try:
            df = pd.read_csv(csv_file, dtype=str)
            patient_count = len(df)
            
            # Extract CAPS name from filename
            caps_name = csv_file.stem.replace('encaminhado__', '').replace('_', ' ')
            
            caps_data.append({
                'caps': caps_name,
                'count': patient_count,
                'filename': csv_file.name
            })
            
            total_patients += patient_count
            
        except Exception as e:
            logger.warning('Erro processando %s: %s', csv_file, e)
    
    # Sort by patient count (descending)
    caps_data.sort(key=lambda x: x['count'], reverse=True)
    
    # Generate report
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write("=" * 60 + "\n")
        f.write("RELATÓRIO DE PACIENTES POR CAPS\n")
        f.write(f"Data: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
        f.write("=" * 60 + "\n\n")
        
        f.write(f"TOTAL GERAL: {total_patients} pacientes\n")
        f.write(f"DISTRIBUÍDOS EM: {len(caps_data)} CAPS diferentes\n\n")
        
        f.write("DETALHAMENTO POR CAPS:\n")
        f.write("-" * 60 + "\n")
        
        for i, data in enumerate(caps_data, 1):
            percentage = (data['count'] / total_patients * 100) if total_patients > 0 else 0
            f.write(f"{i:2d}. {data['caps']:<30} {data['count']:>4} pacientes ({percentage:5.1f}%)\n")
        
        f.write("-" * 60 + "\n")
        f.write(f"TOTAL: {total_patients:>39} pacientes (100.0%)\n\n")
        
        f.write("ARQUIVOS GERADOS:\n")
        f.write("-" * 40 + "\n")
        for data in caps_data:
            f.write(f"• {data['filename']}\n")
        
        f.write("\n" + "=" * 60 + "\n")
        f.write("Arquivos localizados em: by_encaminhado_clean/\n")
        f.write("=" * 60 + "\n")
    
    logger.info('Relatório de pacientes criado: %s', report_file)
    logger.info('Total de pacientes: %s', total_patients)
    logger.info('Distribuídos em %s CAPS', len(caps_data))


def generate_charts(df, output_dir: Path) -> None:
    """Gera gráficos PNG com estatísticas básicas do dataset consolidado."""
    def annotate_bar_values(ax) -> None:
        if hasattr(ax, "bar_label"):
            for container in ax.containers:
                ax.bar_label(container, fmt="%d", padding=2)
        else:
            for patch in ax.patches:
                height = patch.get_height()
                if height is None:
                    continue
                ax.annotate(
                    f"{int(height)}",
                    (patch.get_x() + patch.get_width() / 2, height),
                    ha="center",
                    va="bottom",
                    fontsize=8,
                    xytext=(0, 2),
                    textcoords="offset points",
                )

    if df is None or df.empty:
        logger.info('Sem dados para gerar gráficos.')
        return

    charts_dir = output_dir / 'graficos'
    charts_dir.mkdir(parents=True, exist_ok=True)

    date_col = find_column(df, ['dia alta', 'data'])
    if date_col:
        series = pd.to_datetime(df[date_col], errors='coerce', dayfirst=True)
        month_counts = (
            series.dropna()
            .dt.to_period('M')
            .astype(str)
            .value_counts()
            .sort_index()
        )
        if not month_counts.empty:
            plt.figure(figsize=(10, 4))
            ax = month_counts.plot(kind='bar')
            plt.title('Altas por mês')
            plt.xlabel('Mês')
            plt.ylabel('Quantidade')
            annotate_bar_values(ax)
            plt.tight_layout()
            out_file = charts_dir / 'altas_por_mes.png'
            plt.savefig(out_file, dpi=150)
            plt.close()
            logger.info('Gráfico gerado: %s', out_file)
    else:
        logger.info('Coluna de data não encontrada; gráfico "Altas por mês" não gerado.')

    enc_col = find_column(df, ['encaminhado', 'encaminhamento'])
    if enc_col:
        caps = (
            df[enc_col]
            .fillna('')
            .astype(str)
            .str.strip()
            .replace('', 'VAZIO')
        )
        caps_counts = caps.value_counts().sort_values(ascending=False)
        if not caps_counts.empty:
            plt.figure(figsize=(10, 5))
            ax = caps_counts.plot(kind='bar')
            plt.title('Altas por CAPS')
            plt.xlabel('CAPS')
            plt.ylabel('Quantidade')
            plt.xticks(rotation=45, ha='right')
            annotate_bar_values(ax)
            plt.tight_layout()
            out_file = charts_dir / 'altas_por_caps.png'
            plt.savefig(out_file, dpi=150)
            plt.close()
            logger.info('Gráfico gerado: %s', out_file)
    else:
        logger.info('Coluna "Encaminhado" não encontrada; gráfico "Altas por CAPS" não gerado.')

    if date_col and enc_col:
        series = pd.to_datetime(df[date_col], errors='coerce', dayfirst=True)
        months = series.dt.to_period('M').astype(str)
        caps = (
            df[enc_col]
            .fillna('')
            .astype(str)
            .str.strip()
            .replace('', 'VAZIO')
        )

        pivot = pd.crosstab(months, caps)
        if not pivot.empty:
            top_caps = pivot.sum(axis=0).sort_values(ascending=False).head(12).index
            pivot = pivot[top_caps]

            plt.figure(figsize=(10, 6))
            data = pivot.T.values
            plt.imshow(data, aspect='auto', cmap='magma')
            plt.title('Altas por Mês x CAPS (Top 12)')
            plt.xlabel('Mês')
            plt.ylabel('CAPS')
            plt.xticks(range(len(pivot.index)), pivot.index, rotation=45, ha='right')
            plt.yticks(range(len(pivot.columns)), pivot.columns)
            plt.colorbar(label='Quantidade')
            max_val = data.max() if data.size > 0 else 0
            for y in range(data.shape[0]):
                for x in range(data.shape[1]):
                    val = int(data[y, x])
                    color = 'white' if max_val and val >= max_val * 0.5 else 'black'
                    plt.text(x, y, str(val), ha='center', va='center', fontsize=7, color=color)
            plt.tight_layout()
            out_file = charts_dir / 'altas_mes_x_caps.png'
            plt.savefig(out_file, dpi=150)
            plt.close()
            logger.info('Gráfico gerado: %s', out_file)


def create_clean_encaminhado_files(source_dir: Path, dest_dir: Path):
    """Cria versão limpa dos arquivos de encaminhamento, removendo linhas problemáticas."""
    import shutil
    
    if not source_dir.exists():
        return
    
    dest_dir.mkdir(parents=True, exist_ok=True)
    
    for csv_file in source_dir.glob('*.csv'):
        try:
            df = pd.read_csv(csv_file, dtype=str)
            df_clean = clean_dataframe(df)

            dest_file = dest_dir / csv_file.name
            df_clean.to_csv(dest_file, index=False)
            
            logger.info(
                'Arquivo limpo criado: %s (%s -> %s linhas)',
                csv_file.name,
                len(df),
                len(df_clean),
            )
            
        except Exception as e:
            logger.error('Erro processando %s: %s', csv_file, e)
            # Copy original file if cleaning fails
            shutil.copy2(csv_file, dest_dir / csv_file.name)


if __name__ == '__main__':
    main()
